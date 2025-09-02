import os
import sys
import unicodedata
import json
import pandas as pd
from typing import Tuple, List, Optional
from io import BytesIO
import re
from datetime import datetime, time, timedelta
from flask import (
    Flask, render_template, request, redirect,
    url_for, send_file, session
)
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev_secret_key')

# Carpetas de entrada/salida
UPLOAD_FOLDER = os.path.abspath(os.path.dirname(__file__))
OUTPUT_FOLDER = os.path.join(UPLOAD_FOLDER, 'outputs')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Constantes de negocio
IDEAL_PER_LEADER = 21
ASSIGN_WINDOW    = 2

# Campos a copiar cuando se cruzan nóminas
FILL_FROM_DONOR = ['USUARIO', 'CONTRATO', 'MODALIDAD']

# =========================
# Helpers
# =========================
def to_time(val):
    if pd.isna(val):
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(val.strip(), fmt).time()
            except:
                pass
    if isinstance(val, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(val))).time()
        except:
            pass
    return None

def normalize(text):
    s = str(text).strip().lower()
    return unicodedata.normalize('NFKD', s)

def norm_dni(x) -> str:
    """Normaliza DNI a solo dígitos (quita puntos y '.0')."""
    return ''.join(ch for ch in str(x) if ch.isdigit())

# --- normalizaciones de contenido ---
def norm_usuario(x) -> str:
    """Normaliza usuario a 'u########'. Acepta números, 'U 400769', '400769.0', etc."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ''
    s = str(x).strip()
    if not s:
        return ''
    s = s.replace(' ', '').replace('-', '').replace('_', '')
    if s.lower().startswith('u'):
        digits = ''.join(ch for ch in s[1:] if ch.isdigit())
        return f'u{digits}' if digits else ''
    try:
        n = float(s)
        if n.is_integer():
            return f'u{int(n)}'
    except:
        pass
    digits = ''.join(ch for ch in s if ch.isdigit())
    return f'u{digits}' if digits else s.lower()

def norm_activo(x: str) -> str:
    return str(x).strip().upper()

def norm_contrato(x: str) -> str:
    """
    Devuelve solo '30 hs' o '36 hs' (u otro número si aparece),
    eliminando palabras como 'Temporal' o 'Definitivo' y nombres de meses.
    """
    s = str(x or '').strip()
    if not s:
        return ''

    u = s.upper()

    # quitar meses y palabras no deseadas
    remove_tokens = [
        'ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO',
        'AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE',
        'TEMPORAL','DEFINITIVO','TEMP','DEF'
    ]
    for t in remove_tokens:
        u = u.replace(t, ' ')

    # buscar horas: primero patrones con HS, luego números sueltos (30 o 36)
    m = re.search(r'(\d{2})\s*H?S', u)
    if not m:
        m = re.search(r'\b(30|36)\b', u)

    if m:
        horas = int(m.group(1))
        return f'{horas} hs'

    # fallback: cualquier número
    m = re.search(r'(\d+)', u)
    if m:
        return f"{int(m.group(1))} hs"

    return ''

def rename_to_canon(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        'DNI'      : ['dni', 'documento', 'doc', 'id'],
        'USUARIO'  : ['usuario', 'user', 'legajo', 'employee id', 'usuario sap'],
        'NOMBRE'   : ['nombre', 'name'],
        'SUPERIOR' : ['superior', 'supervisor', 'líder', 'lider', 'coordinador'],
        'SERVICIO' : ['servicio', 'skill', 'servicio/skill', 'segmento'],
        'INGRESO'  : ['ingreso', 'hora ingreso', 'ing hora', 'inicio', 'entrada', 'horario', 'horarios'],
        'ACTIVO'   : ['activo', 'estado'],
        'JEFE'     : ['jefe', 'jefatura', 'manager'],
        'CONTRATO' : ['contrato', 'contrat'],
        'MODALIDAD': ['modalidad', 'home / presencial', 'home/presencial', 'home', 'presencial'],
    }
    lower_map = {c.strip().lower(): c for c in df.columns}
    mapping = {}
    for canon, opts in aliases.items():
        if canon.lower() in lower_map:
            mapping[lower_map[canon.lower()]] = canon
            continue
        for opt in opts:
            if opt in lower_map:
                mapping[lower_map[opt]] = canon
                break
    if mapping:
        df = df.rename(columns=mapping)
    return df

def _canon(df: pd.DataFrame) -> pd.DataFrame:
    """Renombra a canónico, quita columnas duplicadas y asegura índice único."""
    d = df.copy()
    d.columns = d.columns.map(lambda c: str(c).strip())
    d = rename_to_canon(d)
    d = d.loc[:, ~pd.Index(d.columns).duplicated(keep='first')].copy()
    if d.index.has_duplicates:
        d = d.reset_index(drop=True)
    return d


# === Añade cerca de los helpers ===
def choose_join_key(df_full: pd.DataFrame, df_part: pd.DataFrame) -> Optional[str]:
    """Elige la mejor clave disponible para cruzar."""
    for col in ('DNI', 'NOMBRE', 'USUARIO'):
        if col in df_full.columns and col in df_part.columns:
            return col
    return None

def filter_by_boss_or_all(df: pd.DataFrame) -> pd.DataFrame:
    boss_env = os.getenv('BOSS_NAME', 'ANGEL AGUSTIN ROMERO').strip().lower()
    if 'JEFE' in df.columns:
        mask = df['JEFE'].apply(normalize) == boss_env
        if mask.any():
            return df[mask]
    return df

def is_rrss(val) -> bool:
    if pd.isna(val):
        return False
    s = normalize(val)
    return ('rrss' in s) or ('redes sociales' in s) or ('social media' in s) or ('soporte rrss' in s)

def pick_nomina_sheet(xls: pd.ExcelFile) -> str:
    names = [n.strip() for n in xls.sheet_names]
    preferred = [
        'NOMINA SEPTIEMBRE', 'NOMINA OCTUBRE', 'NOMINA AGOSTO',
        'NOMINA JULIO', 'NOMINA JUNIO', 'NOMINA MAYO'
    ]
    for p in preferred:
        for n in names:
            if n.lower() == p.lower():
                return n
    for n in names:
        ln = n.lower()
        if ln.startswith('nomina') and 'rrss' not in ln:
            return n
    return names[0] if names else None

def load_nomina(filename: str, forced_sheet: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame, str, List[str]]:
    path = os.path.join(UPLOAD_FOLDER, filename)
    selected_sheet = None
    sheet_list: List[str] = []
    if filename.lower().endswith(('.xls', '.xlsx')):
        xls = pd.ExcelFile(path)
        sheet_list = xls.sheet_names
        selected_sheet = forced_sheet if (forced_sheet in sheet_list) else pick_nomina_sheet(xls)
        df_src = pd.read_excel(xls, sheet_name=selected_sheet)
    else:
        df_src = pd.read_csv(path)

    df_original_cols = df_src.copy()
    df = df_src.copy()
    df.columns = df.columns.map(lambda c: str(c).strip())
    df = rename_to_canon(df)
    df = filter_by_boss_or_all(df)
    if 'SERVICIO' in df.columns:
        df = df[~df['SERVICIO'].apply(is_rrss)]
    required = ['SUPERIOR','NOMBRE','INGRESO','SERVICIO','ACTIVO']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")
    df = df.dropna(subset=['SUPERIOR']).copy()
    df['ING_TIME'] = df['INGRESO'].apply(to_time)
    return df, df_original_cols, (selected_sheet or ''), sheet_list

# ======= copiar “formato de nómina” (anchos, freeze panes) =======
def clone_layout_from_source(src_path: str, src_sheet: str, dest_ws):
    try:
        wb_src = load_workbook(src_path)
        if src_sheet not in wb_src.sheetnames:
            return
        ws_src = wb_src[src_sheet]
        dest_ws.freeze_panes = ws_src.freeze_panes
        for col_letter, dim in ws_src.column_dimensions.items():
            if dim.width:
                dest_ws.column_dimensions[col_letter].width = dim.width
    except Exception:
        pass

def write_df_to_ws_preserving_nomina(ws, df_out: pd.DataFrame, number_formats=None):
    for j, col in enumerate(df_out.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    for i, row in enumerate(df_out.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if number_formats and df_out.columns[j-1] in number_formats:
                fmt = number_formats[df_out.columns[j-1]]
                if isinstance(val, time):
                    cell.value = datetime.combine(datetime(2000,1,1), val)
                    cell.number_format = fmt
                elif isinstance(val, str):
                    try:
                        t = datetime.strptime(val.strip(), "%H:%M").time()
                        cell.value = datetime.combine(datetime(2000,1,1), t)
                        cell.number_format = fmt
                    except:
                        pass

def ensure_readable_widths(ws, widths_by_header: dict):
    from openpyxl.utils import get_column_letter
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None:
            continue
        headers[str(val).strip()] = col_idx
    for header, width in widths_by_header.items():
        col_idx = headers.get(header)
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        current = ws.column_dimensions[col_letter].width
        if current is None or current < float(width):
            ws.column_dimensions[col_letter].width = float(width)

# ======= armado en ORDEN FIJO para distribución =======
def build_nomina_export(df_src: pd.DataFrame, df_norm: pd.DataFrame, rep2leader: dict) -> pd.DataFrame:
    src = df_src.copy()
    src.columns = src.columns.map(lambda c: str(c).strip())
    src_canon = rename_to_canon(src.copy())
    if 'NOMBRE' in src_canon.columns:
        mask = src_canon['NOMBRE'].astype(str).isin(rep2leader.keys())
        src_canon = src_canon[mask].copy()
    else:
        src_canon = df_norm.copy()
    canon_to_out = {
        'DNI'      : 'DNI',
        'USUARIO'  : 'USUARIO',
        'NOMBRE'   : 'Nombre',
        'SUPERIOR' : 'Superior',
        'SERVICIO' : 'Servicio',
        'INGRESO'  : 'Ingreso',
        'ACTIVO'   : 'Estado',
        'CONTRATO' : 'CONTRATO',
        'MODALIDAD': 'MODALIDAD',
        'JEFE'     : 'JEFE',
    }
    desired_order = ['DNI','USUARIO','Nombre','Superior','Servicio','Ingreso','Estado','CONTRATO','MODALIDAD','JEFE']
    cols_data = {}
    for canon, out_name in canon_to_out.items():
        if canon in src_canon.columns:
            cols_data[out_name] = src_canon[canon].values
        else:
            cols_data[out_name] = [''] * len(src_canon)
    df_out = pd.DataFrame(cols_data, index=src_canon.index)[desired_order]
    if 'NOMBRE' in src_canon.columns:
        nombres = src_canon['NOMBRE'].astype(str)
        df_out['ASIGNADO_A'] = nombres.map(rep2leader).fillna('')
    else:
        df_out['ASIGNADO_A'] = ''
    return df_out

# ========= CRUCE (completar incompleta con completa) =========
def read_nomina_raw(filename: str, forced_sheet: Optional[str] = None) -> Tuple[pd.DataFrame, str]:
    path = os.path.join(UPLOAD_FOLDER, filename)
    if filename.lower().endswith(('.xls', '.xlsx')):
        xls = pd.ExcelFile(path)
        sheet = forced_sheet if (forced_sheet in xls.sheet_names) else pick_nomina_sheet(xls)
        df = pd.read_excel(xls, sheet_name=sheet)
        df.columns = df.columns.map(lambda c: str(c).strip())
        return df, sheet
    else:
        df = pd.read_csv(path)
        df.columns = df.columns.map(lambda c: str(c).strip())
        return df, ''

# Donante: usar TODAS las hojas NOMINA del libro (más recientes primero)
MONTHS = {'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
          'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12}

def _sheet_rank(name: str):
    import re
    n = name.upper()
    ys = [int(x) for x in re.findall(r'(\d{2,4})', n)]
    year = max(ys) if ys else 999
    month = 0
    for k, v in MONTHS.items():
        if k in n:
            month = v; break
    return (year, month)

def read_nomina_all_sheets(filename: str) -> pd.DataFrame:
    path = os.path.join(UPLOAD_FOLDER, filename)
    xls = pd.ExcelFile(path)
    sheets = [s for s in xls.sheet_names if 'NOMINA' in s.upper() and 'RRSS' not in s.upper()]
    sheets.sort(key=_sheet_rank, reverse=True)  # más recientes primero
    parts = []
    for s in sheets:
        df = pd.read_excel(xls, sheet_name=s)
        df.columns = [str(c).strip() for c in df.columns]
        parts.append(df)
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

def pick_donor_receiver(df_a: pd.DataFrame, df_b: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    a = _canon(df_a); b = _canon(df_b)
    score_a = (a['USUARIO'].notna().sum() if 'USUARIO' in a.columns else 0) \
            + (a['CONTRATO'].notna().sum() if 'CONTRATO' in a.columns else 0)
    score_b = (b['USUARIO'].notna().sum() if 'USUARIO' in b.columns else 0) \
            + (b['CONTRATO'].notna().sum() if 'CONTRATO' in b.columns else 0)
    return (df_a, df_b) if score_a >= score_b else (df_b, df_a)

CROSS_ORDER = ['DNI','USUARIO','Nombre','Superior','Servicio','Ingreso','Estado','CONTRATO','MODALIDAD','JEFE']

def cross_fill_users_contracts(
    df_full_src: pd.DataFrame,
    df_part_src: pd.DataFrame,
    fields: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    COMPLETA la nómina INCOMPLETA con datos de la COMPLETA.
    - Busca por DNI; si no encuentra, fallback por NOMBRE normalizado.
    - Ante duplicados en el donante, toma el registro más reciente.
    - Rellena SOLO vacíos en la receptora (no sobreescribe).
    - Devuelve columnas en CROSS_ORDER.
    """
    if fields is None:
        fields = FILL_FROM_DONOR

    full = _canon(df_full_src)
    part = _canon(df_part_src)

    def make_maps(df: pd.DataFrame, key: str, norm_key, fields: List[str]):
        d = df.copy()
        d = d.loc[:, ~pd.Index(d.columns).duplicated(keep='first')].copy()
        if d.index.has_duplicates:
            d = d.reset_index(drop=True)
        if key not in d.columns:
            return {}
        ser = d[key]
        if isinstance(ser, pd.DataFrame):
            ser = ser.iloc[:, 0]
        d['_k'] = ser.apply(norm_key)
        d = d[~d['_k'].isna() & (d['_k'] != '')]
        d = d.drop_duplicates(subset=['_k'], keep='first')  # más reciente
        maps = {}
        for col in fields:
            if col in d.columns:
                maps[col] = pd.Series(d[col].values, index=d['_k']).to_dict()
        return maps

    has_dni = 'DNI' in full.columns and 'DNI' in part.columns
    maps_by_dni  = make_maps(full, 'DNI',   norm_dni,  fields) if has_dni else {}
    maps_by_name = make_maps(full, 'NOMBRE', normalize, fields) if 'NOMBRE' in full.columns else {}

    # asegurar columnas base en la receptora
    base_cols = ['DNI','USUARIO','NOMBRE','SUPERIOR','SERVICIO','INGRESO','ACTIVO','CONTRATO','MODALIDAD','JEFE']
    for c in base_cols:
        if c not in part.columns:
            part[c] = ''

    # rellenar vacíos con prioridad DNI -> Nombre
    def _fill(row, col):
        val = row.get(col, '')
        if pd.notna(val) and str(val).strip() != '':
            return val
        if maps_by_dni:
            k = norm_dni(row.get('DNI', ''))
            if k in maps_by_dni.get(col, {}):
                return maps_by_dni[col][k]
        if maps_by_name:
            k = normalize(row.get('NOMBRE', ''))
            return maps_by_name.get(col, {}).get(k, val)
        return val

    for col in fields:
        part[col] = part.apply(lambda r: _fill(r, col), axis=1)

    # normalizaciones finales
    if 'USUARIO' in part.columns:
        part['USUARIO'] = part['USUARIO'].apply(norm_usuario)
    if 'ACTIVO' in part.columns:
        part['ACTIVO'] = part['ACTIVO'].apply(norm_activo)
    if 'CONTRATO' in part.columns:
        part['CONTRATO'] = part['CONTRATO'].apply(norm_contrato)
    if 'MODALIDAD' in part.columns:
        part['MODALIDAD'] = part['MODALIDAD'].astype(str).str.strip()

    # salida visible y orden fijo
    part_visible = part.rename(columns={
        'NOMBRE':'Nombre', 'SUPERIOR':'Superior', 'SERVICIO':'Servicio',
        'INGRESO':'Ingreso', 'ACTIVO':'Estado'
    })
    for c in CROSS_ORDER:
        if c not in part_visible.columns:
            part_visible[c] = ''
    return part_visible[CROSS_ORDER].copy()

def make_audit_sheets(df_full_src: pd.DataFrame, df_part_src: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Devuelve (no_match, diffs) para auditoría del cruce, tolerante a columnas faltantes/vacías."""
    full = _canon(df_full_src)
    part = _canon(df_part_src)

    # Si alguno viene vacío, devolvemos auditorías vacías
    if full.empty or part.empty:
        no_match = pd.DataFrame(columns=['DNI', 'Nombre'])
        diffs = pd.DataFrame(columns=['DNI','Nombre','Campo','Receptora','Donante'])
        return no_match, diffs

    # Elegir clave disponible
    key = choose_join_key(full, part)
    if not key:
        # No hay forma de cruzar; todo lo de la receptora queda como "no match"
        nm = part.copy()
        nm['Nombre'] = nm['NOMBRE'] if 'NOMBRE' in nm.columns else ''
        cols = []
        if 'DNI' in nm.columns:
            cols.append('DNI')
        cols.append('Nombre')
        no_match = nm[cols]
        diffs = pd.DataFrame(columns=['DNI','Nombre','Campo','Receptora','Donante'])
        return no_match, diffs

    # Normalizador según clave
    if key == 'DNI':
        norm = norm_dni
    elif key == 'USUARIO':
        norm = norm_usuario
    else:
        norm = normalize

    f = full.copy()
    p = part.copy()
    f['_k'] = f[key].apply(norm)
    p['_k'] = p[key].apply(norm)

    # No match (receptora sin match en donante)
    no_match_keys = set(p['_k']) - set(f['_k'])
    nm = p[p['_k'].isin(no_match_keys)].copy()
    nm['Nombre'] = nm['NOMBRE'] if 'NOMBRE' in nm.columns else ''
    cols = []
    if 'DNI' in nm.columns:
        cols.append('DNI')
    cols.append('Nombre')
    no_match = nm[cols]

    # Diferencias en campos relevantes
    merged = pd.merge(p, f, on='_k', how='inner', suffixes=('_R', '_D'))
    check_cols = [('JEFE','JEFE'), ('SUPERIOR','SUPERIOR'), ('INGRESO','INGRESO'), ('SERVICIO','SERVICIO')]
    rows = []
    for _, r in merged.iterrows():
        nombre = r.get('NOMBRE_R', '')
        dni = r.get('DNI_R', '')
        for cr, cd in check_cols:
            vr = r.get(f'{cr}_R', '')
            vd = r.get(f'{cd}_D', '')
            if cr == 'INGRESO':
                tr = to_time(vr); td = to_time(vd)
                vr = tr.strftime('%H:%M') if tr else str(vr)
                vd = td.strftime('%H:%M') if td else str(vd)
            if str(vr).strip() != str(vd).strip():
                rows.append([dni, nombre, cr, str(vr), str(vd)])
    diffs = pd.DataFrame(rows, columns=['DNI','Nombre','Campo','Receptora','Donante'])
    return no_match, diffs
# =========================
# Rutas
# =========================
@app.route('/', methods=['GET','POST'])
def upload():
    if request.method == 'POST':
        files = request.files.getlist('files')
        if not files or files == [None] or files == []:
            f = request.files.get('file')
            if not f:
                return render_template('upload.html', error='Debes subir uno o dos archivos')
            files = [f]
        saved = []
        for f in files[:2]:
            fn = f.filename
            if not fn:
                continue
            path = os.path.join(UPLOAD_FOLDER, fn)
            f.save(path)
            saved.append(fn)
        if len(saved) == 1:
            return redirect(url_for('select', filename=saved[0]))
        elif len(saved) >= 2:
            session['cross_files'] = saved[:2]
            return redirect(url_for('cross'))
        else:
            return render_template('upload.html', error='No se recibió ningún archivo válido.')
    return render_template('upload.html')

@app.route('/cross', methods=['GET'])
def cross():
    files = session.get('cross_files', [])
    if len(files) < 2:
        return redirect(url_for('upload'))

    fn_a, fn_b = files[0], files[1]
    df_a, sheet_a = read_nomina_raw(fn_a)
    df_b, sheet_b = read_nomina_raw(fn_b)

    donor_src, recv_src = pick_donor_receiver(df_a, df_b)
    donor_name = fn_a if donor_src is df_a else fn_b
    recv_name  = fn_b if donor_src is df_a else fn_a
    recv_sheet = sheet_b if donor_src is df_a else sheet_a

    # DONANTE: usar TODAS las hojas NOMINA del libro (más recientes primero)
    donor_pool = read_nomina_all_sheets(donor_name) if donor_name.lower().endswith(('.xls', '.xlsx')) else donor_src
    # Fallback si no se encontraron hojas "NOMINA" o quedó vacío
    if donor_pool is None or getattr(donor_pool, 'empty', True):
        donor_pool = _canon(donor_src)
    # salida principal (receptora completada)
    df_out = cross_fill_users_contracts(donor_pool, recv_src)

    # auditoría contra el POOL
    no_match, diffs = make_audit_sheets(donor_pool, recv_src)

    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=(recv_sheet or 'NOMINA'))

    src_path = os.path.join(UPLOAD_FOLDER, recv_name)
    clone_layout_from_source(src_path, recv_sheet, ws)

    write_df_to_ws_preserving_nomina(ws, df_out, number_formats={'Ingreso': 'hh:mm'})
    try:
        ensure_readable_widths(ws, {
            'DNI': 12, 'USUARIO': 14, 'Nombre': 22, 'Superior': 24,
            'Servicio': 18, 'Ingreso': 10, 'Estado': 12,
            'CONTRATO': 14, 'MODALIDAD': 14, 'JEFE': 22
        })
    except Exception:
        pass

    # hojas de auditoría
    ws_nm = wb.create_sheet(title='No Match')
    if not no_match.empty:
        write_df_to_ws_preserving_nomina(ws_nm, no_match)
    else:
        ws_nm.append(['OK'])

    

    wb.save(output)
    output.seek(0)

    out_name = 'Nomina_cruzada.xlsx'
    with open(os.path.join(OUTPUT_FOLDER, out_name), 'wb') as f:
        f.write(output.getvalue())

    return render_template('cross.html',
                           donor_file=donor_name,
                           receiver_file=recv_name,
                           file_name=out_name)

@app.route('/select/<filename>', methods=['GET','POST'])
def select(filename):
    forced_sheet = request.args.get('sheet') if request.method == 'GET' else request.form.get('sheet')
    try:
        df, df_src, selected_sheet, all_sheets = load_nomina(filename, forced_sheet=forced_sheet)
    except Exception as e:
        return render_template('upload.html', error=f"Error leyendo la nómina: {e}")

    leaders      = sorted(df['SUPERIOR'].dropna().astype(str).unique())
    time_options = [f"{h:02d}:00" for h in range(24)]
    all_services = sorted(df['SERVICIO'].dropna().astype(str).unique())

    if request.method == 'POST':
        start_times   = {L: request.form.get(f'start_{L}') for L in leaders}
        positions     = {L: request.form.get(f'position_{L}') for L in leaders}
        leader_skills = {L: request.form.getlist(f'skill_{L}') for L in leaders}
        for L, lst in list(leader_skills.items()):
            if not lst or lst == ['']:
                leader_skills[L] = []

        windows = {}; windows_str = {}
        for L, st in start_times.items():
            if not st: continue
            t0 = datetime.strptime(st,'%H:%M').time()
            end = (datetime.combine(datetime.today(), t0) + timedelta(hours=ASSIGN_WINDOW)).time()
            windows[L] = (t0, end)
            windows_str[L] = (st, end.strftime('%H:%M'))

        assignments = {L: [] for L in leaders}
        counts      = {L: 0 for L in leaders}
        for _, row in df.sort_values('ING_TIME').iterrows():
            rt  = row['ING_TIME']
            if rt is None: continue
            svc = str(row['SERVICIO'])
            cands = []
            for L, ab in windows.items():
                if not ab: continue
                a, b = ab
                in_window = (a <= rt <= b) if (a <= b) else (rt >= a or rt <= b)
                if not in_window: continue
                sk_list = leader_skills.get(L, [])
                if sk_list:
                    if svc in sk_list: cands.append(L)
                else:
                    cands.append(L)
            if not cands: continue
            chosen = min(cands, key=lambda L: counts[L])
            assignments[chosen].append({
                'rep'     : row['NOMBRE'],
                'service' : svc,
                'ingreso' : rt.strftime('%H:%M'),
                'status'  : row['ACTIVO']
            })
            counts[chosen] += 1

        session['assignments']    = assignments
        session['windows']        = windows_str
        session['positions']      = positions
        session['skills']         = leader_skills
        session['filename']       = filename
        session['selected_sheet'] = selected_sheet
        session['src_cols']       = list(df_src.columns)

        under = [L for L,c in counts.items() if c < IDEAL_PER_LEADER]
        pushers   = [L for L in under if positions.get(L) == 'Pusher']
        referents = [L for L in under if positions.get(L) == 'Referente de Líder']
        leaders_u = [L for L in under if positions.get(L) == 'Líder']
        if pushers:      p = min(pushers, key=lambda L: counts[L])
        elif referents:  p = min(referents, key=lambda L: counts[L])
        elif leaders_u:  p = min(leaders_u, key=lambda L: counts[L])
        else:            p = None

        rep2leader = {}
        for L, infos in assignments.items():
            for info in infos:
                rep2leader[info['rep']] = L

        df_out = build_nomina_export(df_src, df, rep2leader)

        output = BytesIO()
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet(title=(selected_sheet or 'NOMINA'))
        src_path = os.path.join(UPLOAD_FOLDER, filename)
        clone_layout_from_source(src_path, selected_sheet, ws)
        write_df_to_ws_preserving_nomina(ws, df_out, number_formats={'Ingreso': 'hh:mm'})
        try:
            ensure_readable_widths(ws, {
                'DNI': 12, 'USUARIO': 14, 'Nombre': 22, 'Superior': 24,
                'Servicio': 18, 'Ingreso': 10, 'Estado': 12,
                'CONTRATO': 14, 'MODALIDAD': 14, 'JEFE': 22, 'ASIGNADO_A': 20
            })
        except Exception:
            pass

        ws2 = wb.create_sheet(title='Resumen')
        ws2.append(['Líder', 'Ideal', 'Asignados', 'Diferencia'])
        for L in leaders:
            cnt = len(assignments[L])
            ws2.append([L, IDEAL_PER_LEADER, cnt, cnt - IDEAL_PER_LEADER])

        output.seek(0)
        wb.save(output); output.seek(0)
        with open(os.path.join(OUTPUT_FOLDER, 'Nomina_asignada.xlsx'), 'wb') as f:
            f.write(output.getvalue())

        ordered = sorted(windows.keys(), key=lambda L: datetime.strptime(start_times[L], '%H:%M'))
        return render_template('results.html',
                               assignments=assignments,
                               ordered_leaders=ordered,
                               file_name='Nomina_asignada.xlsx',
                               positions=positions,
                               skills=leader_skills,
                               pusher=p,
                               ideal=IDEAL_PER_LEADER,
                               filename=filename)

    return render_template('select.html',
                           filename=filename,
                           leaders=leaders,
                           time_options=time_options,
                           services=all_services,
                           sheets=all_sheets,
                           selected_sheet=selected_sheet)

# Reasignar
@app.route('/reassign/', defaults={'filename': None}, methods=['POST'])
@app.route('/reassign/<path:filename>', methods=['POST'])
def reassign(filename):
    raw        = request.form.get('reassignments', '{}')
    new_assign = json.loads(raw)

    # 1) Usa el filename de la ruta; si falta, cae a session.
    if not filename:
        filename = session.get('filename')

    if not filename:
        return render_template('upload.html',
                               error='La sesión expiró o es inválida. Volvé a subir el archivo para reasignar.')

    forced_sheet = session.get('selected_sheet')  # puede ser None, es OK

    try:
        df, df_src, selected_sheet, _ = load_nomina(filename, forced_sheet=forced_sheet)
    except Exception:
        # 2) NO uses selected_sheet si falló antes de definirse
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = (forced_sheet or 'NOMINA')
        wb.save(output); output.seek(0)
        return send_file(output, as_attachment=True, download_name='Nomina_asignada.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    


    rep2leader = {}
    lookup = df.set_index('NOMBRE').to_dict('index')
    for L, reps in new_assign.items():
        for rep in reps:
            if rep in lookup:
                rep2leader[rep] = L

    df_out = build_nomina_export(df_src, df, rep2leader)

    output = BytesIO(); wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet(title=(selected_sheet or 'NOMINA'))
    src_path = os.path.join(UPLOAD_FOLDER, filename)
    clone_layout_from_source(src_path, selected_sheet, ws)
    write_df_to_ws_preserving_nomina(ws, df_out, number_formats={'Ingreso': 'hh:mm'})
    try:
        ensure_readable_widths(ws, {
            'DNI': 12, 'USUARIO': 14, 'Nombre': 22, 'Superior': 24,
            'Servicio': 18, 'Ingreso': 10, 'Estado': 12,
            'CONTRATO': 14, 'MODALIDAD': 14, 'JEFE': 22, 'ASIGNADO_A': 20
        })
    except Exception:
        pass

    from collections import Counter
    c = Counter(rep2leader.values())
    ws2 = wb.create_sheet(title='Resumen')
    ws2.append(['Líder', 'Ideal', 'Asignados', 'Diferencia'])
    for L, cnt in sorted(c.items()):
        ws2.append([L, IDEAL_PER_LEADER, cnt, cnt - IDEAL_PER_LEADER])

    wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name='Nomina_asignada.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download/<file_name>')
def download(file_name):
    return send_file(os.path.join(OUTPUT_FOLDER, file_name),
                     as_attachment=True,
                     download_name=file_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    print(f"Python: {sys.version}")
    print(">>> Iniciando Flask en http://127.0.0.1:5000 (CTRL+C para salir)")
    if os.name == "nt":
        try:
            import asyncio
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        except Exception:
            pass
    app.run(host='127.0.0.1', port=5000, debug=True)
